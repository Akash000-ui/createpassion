"""
Management command: import_income
Usage: python manage.py import_income --file income.xlsx

Reads an Excel sheet with columns: ID, Income
Updates the income field for matching users by member_id.
Skips rows with missing/invalid data and continues.
"""
import os
from decimal import Decimal, InvalidOperation
from django.core.management.base import BaseCommand
import openpyxl

from mainapp.models import UserProfile


class Command(BaseCommand):
    help = 'Import/update member income from an Excel file (columns: ID, Income)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', default='income.xlsx',
            help='Path to the Excel file (default: income.xlsx)'
        )

    def handle(self, *args, **options):
        filepath = options['file']
        if not os.path.isabs(filepath):
            filepath = os.path.join(os.getcwd(), filepath)

        if not os.path.exists(filepath):
            self.stderr.write(self.style.ERROR(f'File not found: {filepath}'))
            return

        self.stdout.write(f'Loading: {filepath}')
        wb   = openpyxl.load_workbook(filepath)
        ws   = wb.active

        # Detect header row — find ID and Income columns (case-insensitive)
        headers = [str(c.value).strip().upper() if c.value else '' for c in ws[1]]
        try:
            id_col     = headers.index('ID')
            income_col = headers.index('INCOME')
        except ValueError:
            self.stderr.write(self.style.ERROR(
                f'Sheet must have columns "ID" and "Income". Found: {headers}'
            ))
            return

        self.stdout.write(f'Columns — ID: {id_col}, Income: {income_col}')

        updated  = 0
        not_found = 0
        skipped  = 0

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                member_id  = str(row[id_col]).strip() if row[id_col] is not None else ''
                income_raw = row[income_col]

                if not member_id or member_id.upper() == 'NONE':
                    skipped += 1
                    continue

                try:
                    income_val = Decimal(str(income_raw)).quantize(Decimal('0.01'))
                except (InvalidOperation, TypeError):
                    self.stdout.write(self.style.WARNING(
                        f'  Row {i}: Invalid income "{income_raw}" for {member_id} — skipped'
                    ))
                    skipped += 1
                    continue

                rows_updated = UserProfile.objects.filter(member_id=member_id).update(income=income_val)
                if rows_updated:
                    updated += 1
                else:
                    self.stdout.write(f'  Row {i}: Member ID "{member_id}" not found — skipped')
                    not_found += 1

            except Exception as e:
                self.stdout.write(self.style.WARNING(f'  Row {i}: Error — {e}'))
                skipped += 1
                continue

        self.stdout.write(self.style.SUCCESS(
            f'\n✓ Done: {updated} updated, {not_found} ID not found, {skipped} skipped'
        ))
