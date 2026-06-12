from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Sum
from django.utils import timezone
from mainapp.utils.common_utils import login_required_admin, login_required_user, paginate_queryset
from mainapp.utils.validators import validate_image_file
from mainapp.models import UserProfile, Order, KYC, BankDetails, WalletBalance
import openpyxl
from decimal import Decimal, InvalidOperation


def _get_tree_user_ids(root_user):
    """Return a set of UserProfile IDs for root_user and all descendants (BFS)."""
    ids   = set()
    queue = [root_user.id]
    while queue:
        batch = queue[:200]
        queue = queue[200:]
        ids.update(batch)
        children = list(
            UserProfile.objects
            .filter(referred_by_id__in=batch)
            .values_list('id', flat=True)
        )
        queue.extend(children)
    return ids


# ─── User Dashboard ───────────────────────────────────────────────────────────

@login_required_user
def user_dashboard(request):
    user   = UserProfile.objects.get(id=request.session['user_id'])
    orders = Order.objects.filter(user=user).order_by('-order_date')[:5]
    wallet = WalletBalance.objects.filter(user=user).first()
    kyc    = KYC.objects.filter(user=user).first()
    bank   = BankDetails.objects.filter(user=user).first()

    # Tree sales
    now        = timezone.now()
    tree_ids   = _get_tree_user_ids(user)
    tree_qs    = Order.objects.filter(user_id__in=tree_ids).exclude(status='Cancelled')

    daily_sales = tree_qs.filter(
        order_date__date=now.date()
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    monthly_sales = tree_qs.filter(
        order_date__year=now.year,
        order_date__month=now.month,
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    prev_month      = now.month - 1 if now.month > 1 else 12
    prev_month_year = now.year if now.month > 1 else now.year - 1
    prev_monthly_sales = tree_qs.filter(
        order_date__year=prev_month_year,
        order_date__month=prev_month,
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    return render(request, 'user/user_dashboard.html', {
        'profile':             user,
        'orders':              orders,
        'wallet':              wallet,
        'kyc':                 kyc,
        'bank':                bank,
        'daily_sales':         daily_sales,
        'monthly_sales':       monthly_sales,
        'prev_monthly_sales':  prev_monthly_sales,
    })


@login_required_admin
def manage_users(request):
    query  = request.GET.get('q', '').strip()
    status = request.GET.get('status', '')

    users = UserProfile.objects.filter(is_admin=False).order_by('-joining_date')
    if query:
        users = (
            UserProfile.objects.filter(is_admin=False, first_name__icontains=query) |
            UserProfile.objects.filter(is_admin=False, last_name__icontains=query)  |
            UserProfile.objects.filter(is_admin=False, email__icontains=query)      |
            UserProfile.objects.filter(is_admin=False, mobile__icontains=query)     |
            UserProfile.objects.filter(is_admin=False, member_id__icontains=query)
        ).distinct()
    if status == 'active':
        users = users.filter(is_active=True)
    elif status == 'inactive':
        users = users.filter(is_active=False)

    page_obj = paginate_queryset(request, users, per_page=15)
    return render(request, 'admin/users/manage_users.html', {
        'page_obj': page_obj,
        'query': query,
        'status': status,
    })


@login_required_admin
def view_user(request, user_id):
    user = get_object_or_404(UserProfile, id=user_id, is_admin=False)
    orders  = Order.objects.filter(user=user).order_by('-order_date')[:10]
    kyc     = KYC.objects.filter(user=user).first()
    bank    = BankDetails.objects.filter(user=user).first()
    wallet  = WalletBalance.objects.filter(user=user).first()
    return render(request, 'admin/users/view_user.html', {
        'viewed_user': user,
        'orders': orders,
        'kyc': kyc,
        'bank': bank,
        'wallet': wallet,
    })


@login_required_admin
def toggle_user_status(request, user_id):
    if request.method == 'POST':
        user = get_object_or_404(UserProfile, id=user_id, is_admin=False)
        user.is_active = not user.is_active
        user.save()
        action = 'activated' if user.is_active else 'deactivated'
        messages.success(request, f'User {user.get_full_name()} has been {action}.')
    return redirect('manage_users')


@login_required_admin
def update_user_rank(request, user_id):
    if request.method == 'POST':
        user     = get_object_or_404(UserProfile, id=user_id, is_admin=False)
        new_rank = request.POST.get('rank', '').strip()
        valid    = [r[0] for r in UserProfile.RANK_CHOICES]
        if new_rank not in valid:
            messages.error(request, 'Invalid rank selected.')
        else:
            old_rank   = user.rank
            user.rank  = new_rank
            user.save(update_fields=['rank'])
            messages.success(
                request,
                f'Rank updated: {user.get_full_name()} → {new_rank} '
                f'(was {old_rank or "—"})'
            )
    return redirect('view_user', user_id=user_id)


@login_required_admin
def import_income(request):
    results = None
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Please select an Excel file to upload.')
            return redirect('import_income')

        # Validate extension
        if not excel_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, 'Only .xlsx or .xls files are supported.')
            return redirect('import_income')

        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active

            # Read header row
            headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            try:
                uid_col    = headers.index('user_id')
                income_col = headers.index('income')
            except ValueError:
                messages.error(request, 'Excel must have columns named "user_id" and "income".')
                return redirect('import_income')

            updated = []
            skipped = []
            errors  = []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                raw_uid    = row[uid_col]
                raw_income = row[income_col]

                if raw_uid is None and raw_income is None:
                    continue  # blank row

                uid_str = str(raw_uid).strip() if raw_uid is not None else ''
                if not uid_str:
                    errors.append(f'Row {row_num}: empty user_id, skipped.')
                    continue

                try:
                    income_val = Decimal(str(raw_income)).quantize(Decimal('0.01'))
                except (InvalidOperation, TypeError):
                    errors.append(f'Row {row_num}: invalid income value "{raw_income}" for user_id "{uid_str}", skipped.')
                    continue

                # Try member_id first, then DB id
                user = UserProfile.objects.filter(member_id=uid_str, is_admin=False).first()
                if not user and uid_str.isdigit():
                    user = UserProfile.objects.filter(id=int(uid_str), is_admin=False).first()

                if not user:
                    skipped.append(f'Row {row_num}: user_id "{uid_str}" not found.')
                    continue

                user.income = income_val
                user.save(update_fields=['income'])
                updated.append({'uid': uid_str, 'name': user.get_full_name(), 'income': income_val})

            results = {'updated': updated, 'skipped': skipped, 'errors': errors}

            if updated:
                messages.success(request, f'Successfully updated income for {len(updated)} user(s).')
            if skipped:
                messages.warning(request, f'{len(skipped)} user ID(s) not found in the system.')
            if errors:
                messages.warning(request, f'{len(errors)} row(s) had invalid data and were skipped.')

        except Exception as e:
            messages.error(request, f'Failed to read Excel file: {e}')
            return redirect('import_income')

    return render(request, 'admin/users/import_income.html', {'results': results})


# ─── FA+ Member Registration ──────────────────────────────────────────────────

# Ranks that are allowed to register new members and promote to FC
FA_PLUS_RANKS   = ('FA', 'FEM', 'CEM', 'BH', 'BA')
PROMOTED_RANKS  = ('FC', 'FA', 'FEM', 'CEM', 'BH', 'BA')  # all ranks above RT count as used slot
MAX_FC_LIMIT    = 5


def _require_fa_plus(request):
    """Returns (user, None) if FA+, else (None, redirect_response)."""
    user = UserProfile.objects.get(id=request.session['user_id'])
    if user.rank not in FA_PLUS_RANKS:
        messages.error(request, 'You do not have permission to access this page.')
        return None, redirect('user_dashboard')
    return user, None


def _generate_member_id():
    """Generate next member_id in CP##### sequence."""
    import re
    existing = (
        UserProfile.objects
        .filter(member_id__startswith='CP')
        .values_list('member_id', flat=True)
    )
    max_num = 10000
    for mid in existing:
        m = re.match(r'^CP(\d+)$', mid or '')
        if m:
            max_num = max(max_num, int(m.group(1)))
    return f'CP{max_num + 1}'


@login_required_user
def register_member(request):
    user, err = _require_fa_plus(request)
    if err:
        return err

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        mobile     = request.POST.get('mobile', '').strip() or None
        email      = request.POST.get('email', '').strip().lower() or None
        password   = request.POST.get('password', '').strip()

        if not first_name or not last_name:
            messages.error(request, 'First name and last name are required.')
            return render(request, 'user/register_member.html', {'profile': user})

        if not password or len(password) < 6:
            messages.error(request, 'Password must be at least 6 characters.')
            return render(request, 'user/register_member.html', {'profile': user})

        if mobile and UserProfile.objects.filter(mobile=mobile).exists():
            messages.error(request, 'A user with this mobile number already exists.')
            return render(request, 'user/register_member.html', {'profile': user})

        if email and UserProfile.objects.filter(email=email).exists():
            messages.error(request, 'A user with this email already exists.')
            return render(request, 'user/register_member.html', {'profile': user})

        new_member_id = _generate_member_id()
        new_user = UserProfile(
            first_name  = first_name,
            last_name   = last_name,
            mobile      = mobile,
            email       = email,
            member_id   = new_member_id,
            rank        = 'RT',
            referred_by = user,
        )
        new_user.set_password(password)
        new_user.save()

        messages.success(
            request,
            f'Member registered! {new_user.get_full_name()} — '
            f'Member ID: {new_member_id}, Rank: Retailer. '
            f'They can login with Member ID and the password you set.'
        )
        return redirect('my_referrals')

    return render(request, 'user/register_member.html', {'profile': user})


@login_required_user
def my_referrals(request):
    user, err = _require_fa_plus(request)
    if err:
        return err

    referrals    = user.referrals.all().order_by('-created_at')
    fc_count     = referrals.filter(rank__in=PROMOTED_RANKS).count()
    can_promote  = fc_count < MAX_FC_LIMIT
    slots_left   = max(0, MAX_FC_LIMIT - fc_count)

    return render(request, 'user/my_referrals.html', {
        'profile':    user,
        'referrals':  referrals,
        'fc_count':   fc_count,
        'can_promote': can_promote,
        'slots_left': slots_left,
        'max_fc':     MAX_FC_LIMIT,
    })


@login_required_user
def promote_to_fc(request, referred_user_id):
    if request.method != 'POST':
        return redirect('my_referrals')

    user, err = _require_fa_plus(request)
    if err:
        return err

    # The target must be a direct referral of the current user
    target = get_object_or_404(UserProfile, id=referred_user_id, referred_by=user, is_admin=False)

    fc_count = user.referrals.filter(rank__in=PROMOTED_RANKS).count()
    if fc_count >= MAX_FC_LIMIT:
        messages.error(
            request,
            f'You have already promoted {MAX_FC_LIMIT} members to FC. '
            f'You cannot promote any more.'
        )
        return redirect('my_referrals')

    if target.rank in PROMOTED_RANKS:
        messages.warning(request, f'{target.get_full_name()} already has a promoted rank ({target.rank}).')
        return redirect('my_referrals')

    target.rank = 'FC'
    target.save(update_fields=['rank'])
    messages.success(
        request,
        f'{target.get_full_name()} ({target.member_id}) has been promoted to '
        f'Fashion Consultant! You have used {fc_count + 1}/{MAX_FC_LIMIT} FC slots.'
    )
    return redirect('my_referrals')


# ─── Update Profile ───────────────────────────────────────────────────────────

@login_required_user
def update_profile(request):
    user = UserProfile.objects.get(id=request.session['user_id'])
    if request.method == 'POST':
        first_name  = request.POST.get('first_name', '').strip()
        last_name   = request.POST.get('last_name', '').strip()
        email       = request.POST.get('email', '').strip().lower() or None
        mobile      = request.POST.get('mobile', '').strip()
        dob         = request.POST.get('dob', '').strip() or None
        gender      = request.POST.get('gender', '').strip() or None
        address     = request.POST.get('address', '').strip() or None
        profile_pic = request.FILES.get('profile_pic')

        if not first_name or not last_name:
            messages.error(request, 'First name and last name are required.')
            return render(request, 'user/update_profile.html', {'profile': user})

        # Check email uniqueness (exclude self)
        if email and UserProfile.objects.filter(email=email).exclude(id=user.id).exists():
            messages.error(request, 'This email address is already in use.')
            return render(request, 'user/update_profile.html', {'profile': user})

        # Check mobile uniqueness (exclude self)
        if mobile and UserProfile.objects.filter(mobile=mobile).exclude(id=user.id).exists():
            messages.error(request, 'This mobile number is already in use.')
            return render(request, 'user/update_profile.html', {'profile': user})

        if profile_pic:
            try:
                validate_image_file(profile_pic)
            except Exception as e:
                messages.error(request, str(e))
                return render(request, 'user/update_profile.html', {'profile': user})
            user.profile_pic = profile_pic

        user.first_name = first_name
        user.last_name  = last_name
        if email:
            user.email = email
        if mobile:
            user.mobile = mobile
        user.dob     = dob
        user.gender  = gender
        user.address = address
        user.save()

        # Update session name
        request.session['user_name']  = user.get_full_name()
        request.session['user_email'] = user.email or ''
        messages.success(request, 'Profile updated successfully.')
        return redirect('update_profile')

    return render(request, 'user/update_profile.html', {'profile': user})


# ─── Genealogy Tree ───────────────────────────────────────────────────────────

RANK_LABELS = {
    'FC':  'Fashion Consultant',
    'FA':  'Fashion Associate',
    'FEM': 'Fashion Exec. Manager',
    'CEM': 'Chief Exec. Manager',
    'BH':  'Business Head',
    'BA':  'Business Ambassador',
    'RT':  'Retailer',
}
RANK_PCT = {'FC': 21, 'FA': 30, 'FEM': 39, 'CEM': 42, 'BH': 43, 'BA': 44, 'RT': 11}


def _user_to_dict(u):
    if isinstance(u, dict):
        first = u.get('first_name', '')
        last  = u.get('last_name', '')
        uid   = u['id']
        mid   = u.get('member_id') or '—'
        rank  = u.get('rank') or 'BH'
        direct_count = UserProfile.objects.filter(referred_by_id=uid).count()
    else:
        first = u.first_name
        last  = u.last_name
        uid   = u.id
        mid   = u.member_id or '—'
        rank  = u.rank or 'BH'
        direct_count = u.referrals.count()
    return {
        'id':           uid,
        'name':         f'{first} {last}',
        'initials':     (first or '?')[0].upper(),
        'member_id':    mid,
        'rank':         rank,
        'rank_label':   RANK_LABELS.get(rank, rank),
        'commission':   RANK_PCT.get(rank),
        'direct_count': direct_count,
    }


@login_required_user
def genealogy_tree(request, target_id=None):
    current_user = UserProfile.objects.get(id=request.session['user_id'])

    # Determine whose tree to show
    if target_id:
        try:
            target = UserProfile.objects.get(id=target_id, is_admin=False)
        except UserProfile.DoesNotExist:
            target = current_user
    else:
        target = current_user

    root     = _user_to_dict(target)
    children = [_user_to_dict(c) for c in target.referrals.order_by('first_name')]

    # Build ancestry chain (for breadcrumb navigation, up to 10 levels up)
    ancestry = []
    node = target
    for _ in range(10):
        if not node.referred_by_id:
            break
        parent = node.referred_by
        ancestry.insert(0, {
            'id':        parent.id,
            'name':      parent.get_full_name(),
            'member_id': parent.member_id or '—',
        })
        node = parent

    return render(request, 'user/genealogy_tree.html', {
        'profile':      current_user,
        'root':         root,
        'children':     children,
        'ancestry':     ancestry,
        'is_own_tree':  target.id == current_user.id,
    })


@login_required_user
def genealogy_children(request, user_id):
    """AJAX: return direct children of a given user as JSON."""
    from django.http import JsonResponse
    try:
        node     = UserProfile.objects.get(id=user_id, is_admin=False)
        children = [_user_to_dict(c) for c in node.referrals.order_by('first_name')]
        return JsonResponse({'children': children})
    except UserProfile.DoesNotExist:
        return JsonResponse({'children': []})



